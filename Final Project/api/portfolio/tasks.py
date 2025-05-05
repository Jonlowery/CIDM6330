# portfolio/tasks.py (Add CPR handling to Security import)

import os
import logging
import time
import openpyxl
from celery import shared_task, chain
from django.db import transaction, IntegrityError, OperationalError
from django.conf import settings
from django.core.mail import send_mail
from django.utils import timezone
from datetime import date, datetime
# Import models from the current app, including new ones
from .models import (
    Security, Customer, Portfolio, CustomerHolding, MunicipalOffering,
    Salesperson, SecurityType, InterestSchedule # New models
)
# Import Decimal for data cleaning
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
# Import FieldError for specific error handling
from django.core.exceptions import FieldError
# Import validator for email
from django.core.validators import validate_email
from django.core.exceptions import ValidationError as DjangoValidationError


# Setup logging
log = logging.getLogger(__name__)

# --- Helper Functions (Keep existing ones: clean_decimal, clean_date, clean_boolean_from_char) ---

def clean_decimal(value, default=None, max_digits=None, decimal_places=None, non_negative=False):
    """
    Safely convert value to Decimal, applying validation rules.
    Returns default if conversion fails or validation rules are not met.
    """
    if value is None or str(value).strip() == '':
        return default
    try:
        str_value = str(value).strip().replace(',', '') # Handle commas
        if '%' in str_value: str_value = str_value.replace('%', '').strip() # Handle percentage signs
        if str_value.startswith('(') and str_value.endswith(')'): str_value = '-' + str_value[1:-1] # Handle accounting negatives

        d = Decimal(str_value)

        # Validation: Non-negative
        if non_negative and d < 0:
            log.warning(f"Validation failed: Value '{value}' converted to '{d}' is negative, but non-negative required. Using default '{default}'.")
            return default

        # Validation: Max digits and decimal places (if specified)
        # Note: This validation is tricky with Decimal, often better handled at DB level.
        # Basic check for number of decimal places after conversion:
        if decimal_places is not None:
            # Quantize to check/enforce decimal places if needed, though DB handles storage.
            # Be careful quantizing here, might lose precision intended for DB.
            # Let's just check the number of places in the string representation for logging.
            if '.' in str_value:
                actual_dp = len(str_value.split('.')[-1])
                if actual_dp > decimal_places:
                     log.warning(f"Value '{value}' has more than {decimal_places} decimal places. Precision might be lost on save.")
                     # Optional: Quantize here if strict enforcement needed before DB save
                     # d = d.quantize(Decimal('1e-' + str(decimal_places)), rounding=ROUND_HALF_UP)


        # Add check for total digits if max_digits is provided (more complex)
        # For simplicity, relying on DB constraints for max_digits.

        return d
    except (InvalidOperation, TypeError, ValueError):
        log.warning(f"Could not convert '{value}' to Decimal, using default '{default}'")
        return default

def clean_date(value, default=None, date_format='%m/%d/%Y'):
    """
    Safely convert value to Date using a specific format, return default if conversion fails.
    Handles datetime objects from openpyxl.
    Returns None for clearly invalid date strings like '01/  /'.
    """
    if value is None: return default
    if isinstance(value, datetime): return value.date()
    if isinstance(value, date): return value # Already a date object

    # Handle Excel serial dates (numbers)
    if isinstance(value, (int, float)):
        try:
            # Ensure openpyxl is available for this conversion
            from openpyxl.utils.datetime import from_excel
            # Basic check to avoid converting huge numbers that aren't dates
            if value > 2958465: # Heuristic for Excel dates (approx year 9999)
                 log.warning(f"Excel number '{value}' too large, likely not a date. Skipping conversion.")
                 return default
            dt_value = from_excel(value)
            return dt_value.date()
        except ImportError:
            log.error("openpyxl is required for converting Excel numeric dates but not installed.")
            return default
        except (ValueError, TypeError, OverflowError) as e:
             log.warning(f"Could not convert Excel number '{value}' to Date: {e}. Using default '{default}'")
             return default

    # Handle string dates
    if isinstance(value, str):
        value_str = value.strip()
        if not value_str: return default

        # --- Added check for obviously invalid date patterns ---
        if '/' in value_str and value_str.count('/') == 2:
            parts = value_str.split('/')
            if any(not p.strip() for p in parts): # Check for empty parts like in '01/  /'
                log.warning(f"Invalid date string pattern detected: '{value_str}'. Returning None.")
                return None
        # --- End added check ---

        try:
            # Attempt parsing with the primary expected format first
            return datetime.strptime(value_str, date_format).date()
        except (ValueError, TypeError):
            # Try fallback formats if primary fails
            fallback_formats = ('%Y-%m-%d', '%m-%d-%Y', '%Y%m%d')
            for fmt in fallback_formats:
                try: return datetime.strptime(value_str, fmt).date()
                except (ValueError, TypeError): continue
            log.warning(f"Could not parse date string '{value_str}' with formats '{date_format}' or fallbacks {fallback_formats}.")
            return default

    log.warning(f"Value '{value}' type '{type(value)}' could not be converted to Date.")
    return default

def clean_boolean_from_char(value, true_chars=('y', 'yes', 'true', '1', 't')):
    """ Converts a character/string (like 'y'/'n') to Boolean. Case-insensitive. """
    if value is None:
        return None # Or False depending on desired default for missing values
    return str(value).strip().lower() in true_chars

# --- NEW Lookup Import Tasks (Keep existing ones) ---

@shared_task(bind=True, autoretry_for=(OperationalError,), retry_backoff=5, max_retries=3)
def import_salespersons_from_excel(self, file_path):
    """
    Imports or updates Salesperson records from an Excel file.
    Expects columns: 'slsm_id', 'name', and optionally 'email'.
    """
    log.info(f"Starting Salesperson import/update from {file_path}")
    updated_count = 0; created_count = 0; skipped_rows = 0
    max_retries_per_row = 3; retry_delay = 0.5

    # Define expected Excel headers and map to model fields
    header_map = {
        'slsm_id': 'salesperson_id', # Maps to Salesperson model's PK
        'name': 'name',
        'email': 'email', # *** ADDED email mapping ***
    }
    mandatory_internal_names = ['salesperson_id'] # Name and Email are optional in model

    try:
        wb = openpyxl.load_workbook(filename=file_path, data_only=True, read_only=True); ws = wb.active
    except FileNotFoundError: log.error(f"Salesperson file not found: {file_path}"); raise
    except Exception as e: log.error(f"Error opening salesperson file {file_path}: {e}", exc_info=True); raise

    # Read actual headers from file and map them using header_map
    excel_headers = [str(cell.value).strip() if cell.value else None for cell in ws[1]]
    col_idx_map = {}
    processed_excel_headers = []
    for idx, excel_header in enumerate(excel_headers):
        if excel_header:
            normalized_header = excel_header.lower().replace(' ', '_')
            internal_name = header_map.get(normalized_header)
            if internal_name:
                col_idx_map[internal_name] = idx
                processed_excel_headers.append(excel_header)
            else:
                log.warning(f"Salesperson Import: Unrecognized header '{excel_header}' in column {idx+1}. It will be ignored.")

    log.info(f"Salesperson Import: Found and processed Excel headers: {processed_excel_headers}")
    log.info(f"Salesperson Import: Mapped internal names to column indices: {col_idx_map}")

    # Check for mandatory fields
    missing_mandatory = [name for name in mandatory_internal_names if name not in col_idx_map]
    if missing_mandatory:
        log.error(f"Mandatory headers missing from salesperson import file (mapped names: {missing_mandatory}). Cannot proceed.")
        wb.close()
        raise ValueError(f"Mandatory salesperson headers missing: {missing_mandatory}")

    # Iterate through rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw_data = {}
        for internal_name, col_idx in col_idx_map.items():
            if col_idx < len(row): raw_data[internal_name] = row[col_idx]
            else: raw_data[internal_name] = None

        # --- Data Extraction and Cleaning ---
        salesperson_id = raw_data.get('salesperson_id')
        if not salesperson_id: log.warning(f"Salesperson Row {row_idx}: Skip missing slsm_id."); skipped_rows += 1; continue
        salesperson_id = str(salesperson_id).strip() # Keep as string

        name = str(raw_data.get('name', '')).strip() or None # Allow empty name if model allows (blank=True)

        # *** ADDED Email Cleaning ***
        email_raw = raw_data.get('email')
        email = None
        if email_raw:
            email_str = str(email_raw).strip()
            if email_str: # Only proceed if email is not empty after stripping
                try:
                    validate_email(email_str) # Use Django's validator
                    email = email_str
                except DjangoValidationError:
                    log.warning(f"Salesperson Row {row_idx} ID {salesperson_id}: Invalid email format '{email_raw}'. Storing as NULL.")
                    email = None # Set to None if invalid format
        # *** END Email Cleaning ***

        # Prepare defaults dictionary
        data_defaults = {
            'name': name,
            'email': email, # Include cleaned email
            # Set other defaults if needed, e.g., 'is_active': True
        }
        # Remove None values if you don't want to overwrite existing DB values with NULL
        # Decide if you want this behavior or if NULL should overwrite existing values
        # data_defaults = {k: v for k, v in data_defaults.items() if v is not None}

        # --- Database Operation ---
        retries = 0; success = False
        while retries < max_retries_per_row and not success:
            try:
                with transaction.atomic():
                    sp, created = Salesperson.objects.update_or_create(
                        salesperson_id=salesperson_id, # Match based on PK
                        defaults=data_defaults
                    )
                success = True
                if created: created_count += 1; log.debug(f"Salesperson Row {row_idx}: Created: {salesperson_id} (Email: {email})")
                else: updated_count += 1; log.debug(f"Salesperson Row {row_idx}: Updated: {salesperson_id} (Email: {email})")
            except OperationalError as e:
                if 'database is locked' in str(e).lower() and retries < max_retries_per_row - 1:
                    retries += 1; wait_time = retry_delay * (2**retries); log.warning(f"Salesperson Row {row_idx}: DB locked {salesperson_id}. Retry {retries}/{max_retries_per_row-1} in {wait_time:.2f}s"); time.sleep(wait_time)
                else:
                    log.error(f"Salesperson Row {row_idx}: OpError {salesperson_id} retries {retries}: {e}")
                    # Let celery handle final retry based on task decorator
                    if retries >= max_retries_per_row -1: self.retry(exc=e)
                    skipped_rows += 1; break # Break local loop after triggering retry or final failure
            except IntegrityError as e:
                # This could happen if the email being imported conflicts with an existing one
                # AND the email field has unique=True in the model (which it does).
                log.error(f"Salesperson Row {row_idx}: IntegrityError {salesperson_id} (Likely duplicate email '{email}'): {e}");
                skipped_rows += 1; break
            except Exception as e: log.error(f"Salesperson Row {row_idx}: Error {salesperson_id}: {e}", exc_info=True); skipped_rows += 1; break

    wb.close()
    result_message = f"Imported/Updated Salespersons from {os.path.basename(file_path)}. Created: {created_count}, Updated: {updated_count}, Skipped: {skipped_rows}."
    log.info(result_message)
    return file_path # Return path for chaining


@shared_task(bind=True, autoretry_for=(OperationalError,), retry_backoff=5, max_retries=3)
def import_security_types_from_excel(self, file_path):
    """
    Imports or updates SecurityType records from an Excel file.
    Expects columns: 'sec_type', 'meaning'
    NOTE: Name uniqueness constraint removed from model.
    """
    log.info(f"Starting SecurityType import/update from {file_path}")
    updated_count = 0; created_count = 0; skipped_rows = 0
    max_retries_per_row = 3; retry_delay = 0.5

    # Define expected Excel headers and map to model fields
    header_map = {
        'sec_type': 'type_id', # Maps to SecurityType model's PK
        'meaning': 'name', # Maps to the 'name' field
        # Add 'description' if it exists in the Excel file
        # 'description': 'description',
    }
    mandatory_internal_names = ['type_id', 'name'] # Both ID and name are required

    try:
        wb = openpyxl.load_workbook(filename=file_path, data_only=True, read_only=True); ws = wb.active
    except FileNotFoundError: log.error(f"SecurityType file not found: {file_path}"); raise
    except Exception as e: log.error(f"Error opening SecurityType file {file_path}: {e}", exc_info=True); raise

    # Read actual headers and map
    excel_headers = [str(cell.value).strip() if cell.value else None for cell in ws[1]]
    col_idx_map = {}
    processed_excel_headers = []
    for idx, excel_header in enumerate(excel_headers):
        if excel_header:
            normalized_header = excel_header.lower().replace(' ', '_')
            internal_name = header_map.get(normalized_header)
            if internal_name:
                col_idx_map[internal_name] = idx
                processed_excel_headers.append(excel_header)
            else:
                log.warning(f"SecurityType Import: Unrecognized header '{excel_header}' in column {idx+1}. It will be ignored.")

    log.info(f"SecurityType Import: Found and processed Excel headers: {processed_excel_headers}")
    log.info(f"SecurityType Import: Mapped internal names to column indices: {col_idx_map}")

    # Check for mandatory fields
    missing_mandatory = [name for name in mandatory_internal_names if name not in col_idx_map]
    if missing_mandatory:
        log.error(f"Mandatory headers missing from SecurityType import file (mapped names: {missing_mandatory}). Cannot proceed.")
        wb.close()
        raise ValueError(f"Mandatory SecurityType headers missing: {missing_mandatory}")

    # Iterate through rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw_data = {}
        for internal_name, col_idx in col_idx_map.items():
            if col_idx < len(row): raw_data[internal_name] = row[col_idx]
            else: raw_data[internal_name] = None

        # --- Data Extraction and Cleaning ---
        type_id_raw = raw_data.get('type_id')
        type_id = clean_decimal(type_id_raw) # Clean as decimal first
        if type_id is None: log.warning(f"SecurityType Row {row_idx}: Skip missing or invalid sec_type '{type_id_raw}'."); skipped_rows += 1; continue
        try:
            type_id = int(type_id) # Convert to int
        except (ValueError, TypeError):
             log.warning(f"SecurityType Row {row_idx}: Could not convert sec_type '{type_id_raw}' to integer. Skipping."); skipped_rows += 1; continue

        name = str(raw_data.get('name', '')).strip()
        if not name: log.warning(f"SecurityType Row {row_idx} ID {type_id}: Skip missing meaning/name."); skipped_rows += 1; continue

        description = str(raw_data.get('description', '')).strip() or None # Optional description

        # Prepare defaults dictionary
        data_defaults = {
            'name': name,
            'description': description,
        }
        data_defaults = {k: v for k, v in data_defaults.items() if v is not None}

        # --- Database Operation ---
        retries = 0; success = False
        while retries < max_retries_per_row and not success:
            try:
                with transaction.atomic():
                    st, created = SecurityType.objects.update_or_create(
                        type_id=type_id, # Match based on PK
                        defaults=data_defaults
                    )
                success = True
                if created: created_count += 1; log.debug(f"SecurityType Row {row_idx}: Created: {type_id}")
                else: updated_count += 1; log.debug(f"SecurityType Row {row_idx}: Updated: {type_id}")
            except OperationalError as e:
                if 'database is locked' in str(e).lower() and retries < max_retries_per_row - 1:
                    retries += 1; wait_time = retry_delay * (2**retries); log.warning(f"SecurityType Row {row_idx}: DB locked {type_id}. Retry {retries}/{max_retries_per_row-1} in {wait_time:.2f}s"); time.sleep(wait_time)
                else:
                    log.error(f"SecurityType Row {row_idx}: OpError {type_id} retries {retries}: {e}")
                    if retries >= max_retries_per_row -1: self.retry(exc=e)
                    skipped_rows += 1; break
            except IntegrityError as e:
                # This should no longer happen for the name field, but could happen if type_id is duplicated in the file
                log.error(f"SecurityType Row {row_idx}: IntegrityError {type_id}: {e}"); skipped_rows += 1; break
            except Exception as e: log.error(f"SecurityType Row {row_idx}: Error {type_id}: {e}", exc_info=True); skipped_rows += 1; break

    wb.close()
    result_message = f"Imported/Updated SecurityTypes from {os.path.basename(file_path)}. Created: {created_count}, Updated: {updated_count}, Skipped: {skipped_rows}."
    log.info(result_message)
    return file_path # Return path for chaining


@shared_task(bind=True, autoretry_for=(OperationalError,), retry_backoff=5, max_retries=3)
def import_interest_schedules_from_excel(self, file_path):
    """
    Imports or updates InterestSchedule records from an Excel file.
    Expects columns: 'int_sched', 'meaning'
    """
    log.info(f"Starting InterestSchedule import/update from {file_path}")
    updated_count = 0; created_count = 0; skipped_rows = 0
    max_retries_per_row = 3; retry_delay = 0.5

    # Define expected Excel headers and map to model fields
    header_map = {
        'int_sched': 'schedule_code', # Maps to InterestSchedule model's PK
        'meaning': 'name', # Maps to the 'name' field
        # Add other fields if they exist in Excel (e.g., 'ppy_default', 'description')
        # 'ppy_default': 'payments_per_year_default',
        # 'description': 'description',
    }
    mandatory_internal_names = ['schedule_code', 'name'] # Both code and name are required

    try:
        wb = openpyxl.load_workbook(filename=file_path, data_only=True, read_only=True); ws = wb.active
    except FileNotFoundError: log.error(f"InterestSchedule file not found: {file_path}"); raise
    except Exception as e: log.error(f"Error opening InterestSchedule file {file_path}: {e}", exc_info=True); raise

    # Read actual headers and map
    excel_headers = [str(cell.value).strip() if cell.value else None for cell in ws[1]]
    col_idx_map = {}
    processed_excel_headers = []
    for idx, excel_header in enumerate(excel_headers):
        if excel_header:
            normalized_header = excel_header.lower().replace(' ', '_')
            internal_name = header_map.get(normalized_header)
            if internal_name:
                col_idx_map[internal_name] = idx
                processed_excel_headers.append(excel_header)
            else:
                log.warning(f"InterestSchedule Import: Unrecognized header '{excel_header}' in column {idx+1}. It will be ignored.")

    log.info(f"InterestSchedule Import: Found and processed Excel headers: {processed_excel_headers}")
    log.info(f"InterestSchedule Import: Mapped internal names to column indices: {col_idx_map}")

    # Check for mandatory fields
    missing_mandatory = [name for name in mandatory_internal_names if name not in col_idx_map]
    if missing_mandatory:
        log.error(f"Mandatory headers missing from InterestSchedule import file (mapped names: {missing_mandatory}). Cannot proceed.")
        wb.close()
        raise ValueError(f"Mandatory InterestSchedule headers missing: {missing_mandatory}")

    # Iterate through rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw_data = {}
        for internal_name, col_idx in col_idx_map.items():
            if col_idx < len(row): raw_data[internal_name] = row[col_idx]
            else: raw_data[internal_name] = None

        # --- Data Extraction and Cleaning ---
        schedule_code = raw_data.get('schedule_code')
        if not schedule_code: log.warning(f"InterestSchedule Row {row_idx}: Skip missing int_sched."); skipped_rows += 1; continue
        schedule_code = str(schedule_code).strip() # Keep as string

        name = str(raw_data.get('name', '')).strip()
        if not name: log.warning(f"InterestSchedule Row {row_idx} Code {schedule_code}: Skip missing meaning/name."); skipped_rows += 1; continue

        # Optional fields
        ppy_default_raw = raw_data.get('payments_per_year_default')
        ppy_default = clean_decimal(ppy_default_raw)
        ppy_default_int = None
        if ppy_default is not None:
            try:
                 ppy_default_int = int(ppy_default)
                 if ppy_default_int <= 0: # Validate positive
                      log.warning(f"InterestSchedule Row {row_idx} Code {schedule_code}: Invalid ppy_default '{ppy_default_raw}'. Setting to NULL."); ppy_default_int = None
            except (ValueError, TypeError):
                 log.warning(f"InterestSchedule Row {row_idx} Code {schedule_code}: Could not convert ppy_default '{ppy_default_raw}' to integer. Setting to NULL."); ppy_default_int = None

        description = str(raw_data.get('description', '')).strip() or None

        # Prepare defaults dictionary
        data_defaults = {
            'name': name,
            'payments_per_year_default': ppy_default_int,
            'description': description,
        }
        data_defaults = {k: v for k, v in data_defaults.items() if v is not None}

        # --- Database Operation ---
        retries = 0; success = False
        while retries < max_retries_per_row and not success:
            try:
                with transaction.atomic():
                    isc, created = InterestSchedule.objects.update_or_create(
                        schedule_code=schedule_code, # Match based on PK
                        defaults=data_defaults
                    )
                success = True
                if created: created_count += 1; log.debug(f"InterestSchedule Row {row_idx}: Created: {schedule_code}")
                else: updated_count += 1; log.debug(f"InterestSchedule Row {row_idx}: Updated: {schedule_code}")
            except OperationalError as e:
                if 'database is locked' in str(e).lower() and retries < max_retries_per_row - 1:
                    retries += 1; wait_time = retry_delay * (2**retries); log.warning(f"InterestSchedule Row {row_idx}: DB locked {schedule_code}. Retry {retries}/{max_retries_per_row-1} in {wait_time:.2f}s"); time.sleep(wait_time)
                else:
                    log.error(f"InterestSchedule Row {row_idx}: OpError {schedule_code} retries {retries}: {e}")
                    if retries >= max_retries_per_row -1: self.retry(exc=e)
                    skipped_rows += 1; break
            except IntegrityError as e: # Could happen on schedule_code or name (if unique=True wasn't removed from name)
                log.error(f"InterestSchedule Row {row_idx}: IntegrityError {schedule_code}: {e}"); skipped_rows += 1; break
            except Exception as e: log.error(f"InterestSchedule Row {row_idx}: Error {schedule_code}: {e}", exc_info=True); skipped_rows += 1; break

    wb.close()
    result_message = f"Imported/Updated InterestSchedules from {os.path.basename(file_path)}. Created: {created_count}, Updated: {updated_count}, Skipped: {skipped_rows}."
    log.info(result_message)
    return file_path # Return path for chaining

# --- Existing Import Tasks (Keep import_securities, import_customers, import_holdings) ---
# (Code for these tasks remains the same as in the previous version)

@shared_task
def import_securities_from_excel(file_path):
    """
    Imports or updates securities from an Excel file based on CUSIP (sec_id).
    Maps new Excel columns to the revamped Security model.
    (Ensure this task uses the updated Security model fields)
    """
    log.info(f"Starting security import/update from {file_path}")
    updated_count = 0; created_count = 0; skipped_rows = 0
    max_retries_per_row = 3; retry_delay = 0.5

    # Define expected Excel headers based on validation rules
    # Renaming Excel headers to snake_case for internal use
    header_map = {
        'sec_id': 'cusip', # Maps to Security model's PK
        'sec_type': 'security_type_id', # Maps to SecurityType FK (by type_id)
        'mat_dt': 'maturity_date',
        'rate': 'base_rate', # Temporary name for Excel 'rate'
        'sec_desc_1': 'description',
        'issue_dt': 'issue_date',
        'tax_cd': 'tax_code', # Expects 'e' or 't'
        'int_sched': 'interest_schedule_code', # Maps to InterestSchedule FK (by schedule_code)
        'int_day': 'interest_day',
        'int_calc_cd': 'interest_calc_code', # Expects 'a', 'c', or 'h'
        'ppy': 'payments_per_year',
        'prin_paydown': 'prin_paydown_flag', # Temporary name for Excel 'prin_paydown' ('y'/'n')
        'pmt_delay': 'payment_delay_days',
        'rate_dt': 'rate_effective_date', # Optional date for secrate_rate
        'secrate_rate': 'secondary_rate', # Optional override rate
        'factor': 'factor_from_excel', # Temporary name for Excel 'factor'
        # *** ADD CPR MAPPING ***
        # Assuming the Excel column header is 'cpr' (case-insensitive)
        'cpr': 'cpr',
        # ----------------------
        # Add mappings for other potential columns if they exist in the Excel file
        # and map to existing/new Security fields (issuer_name, ratings, sector etc.)
        # Assuming they might exist with similar names for now:
        'issuer_name': 'issuer_name',
        'currency': 'currency',
        'callable': 'callable_flag_excel', # Temporary name for 'y'/'n' or similar
        'moody': 'moody_rating',
        's_and_p': 'sp_rating',
        'fitch': 'fitch_rating',
        'sector': 'sector',
        'issuer_state': 'state_of_issuer',
        'call_date': 'call_date', # Existing optional field
        'wal': 'wal', # Existing optional field
    }

    try:
        wb = openpyxl.load_workbook(filename=file_path, data_only=True, read_only=True); ws = wb.active
    except FileNotFoundError: log.error(f"Security file not found: {file_path}"); raise
    except Exception as e: log.error(f"Error opening security file {file_path}: {e}", exc_info=True); raise

    # Read actual headers from file and map them using header_map
    excel_headers = [str(cell.value).strip() if cell.value else None for cell in ws[1]]
    # Create a map from normalized internal names to column index { 'cusip': 0, 'security_type_id': 1, ... }
    col_idx_map = {}
    processed_excel_headers = [] # Keep track of headers found in the file
    for idx, excel_header in enumerate(excel_headers):
        if excel_header:
            normalized_header = excel_header.lower().replace(' ', '_')
            # Find the internal name from header_map based on the normalized Excel header
            internal_name = None
            for map_key, map_value in header_map.items():
                # Allow flexible matching (e.g., 'sec_id' or 'sec id' in Excel maps to 'cusip')
                if normalized_header == map_key.lower().replace(' ', '_'):
                    internal_name = map_value
                    break
            if internal_name:
                col_idx_map[internal_name] = idx
                processed_excel_headers.append(excel_header)
            else:
                log.warning(f"Security Import: Unrecognized header '{excel_header}' in column {idx+1}. It will be ignored.")

    log.info(f"Security Import: Found and processed Excel headers: {processed_excel_headers}")
    log.info(f"Security Import: Mapped internal names to column indices: {col_idx_map}")

    # Check for mandatory fields based on internal names
    mandatory_internal_names = [
        'cusip', 'maturity_date', 'description', 'issue_date', 'tax_code',
        # 'interest_schedule_code', # FK lookup, not strictly mandatory header if FK can be null
        'interest_day', 'interest_calc_code',
        'payments_per_year', 'prin_paydown_flag', 'payment_delay_days'
        # CPR is optional, so not added here
    ]
    missing_mandatory = [name for name in mandatory_internal_names if name not in col_idx_map]
    if missing_mandatory:
        log.error(f"Mandatory headers missing from security import file (mapped names: {missing_mandatory}). Cannot proceed.")
        wb.close()
        raise ValueError(f"Mandatory security headers missing: {missing_mandatory}")

    # Pre-fetch related objects for efficiency if possible (might be large)
    # Caching within the loop might be more memory-efficient for very large files
    security_types = {st.type_id: st for st in SecurityType.objects.all()}
    interest_schedules = {isc.schedule_code: isc for isc in InterestSchedule.objects.all()}

    # Iterate through rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Create dictionary using internal names and mapped indices
        raw_data = {}
        for internal_name, col_idx in col_idx_map.items():
            if col_idx < len(row): # Ensure index is within row bounds
                raw_data[internal_name] = row[col_idx]
            else:
                raw_data[internal_name] = None # Assign None if column index exceeds row length

        # --- Data Extraction and Cleaning ---
        cusip = raw_data.get('cusip')
        if not cusip: log.warning(f"Sec Row {row_idx}: Skip missing CUSIP."); skipped_rows += 1; continue
        cusip = str(cusip).strip().upper()
        # Basic CUSIP validation (length 9, alphanumeric) - can enhance later
        if len(cusip) != 9 or not cusip.isalnum():
            log.warning(f"Sec Row {row_idx}: Invalid CUSIP format '{raw_data.get('cusip')}'. Skipping row.")
            skipped_rows += 1; continue

        # Foreign Key Lookups (handle missing related objects)
        sec_type_id_raw = raw_data.get('security_type_id')
        sec_type_id = clean_decimal(sec_type_id_raw)
        security_type_instance = None
        if sec_type_id is not None:
            try:
                sec_type_id_int = int(sec_type_id)
                security_type_instance = security_types.get(sec_type_id_int)
                if not security_type_instance:
                    log.warning(f"Sec Row {row_idx} CUSIP {cusip}: SecurityType ID '{sec_type_id_int}' not found in DB. Setting type to NULL.")
            except (ValueError, TypeError):
                 log.warning(f"Sec Row {row_idx} CUSIP {cusip}: Invalid SecurityType ID format '{sec_type_id_raw}'. Setting type to NULL.")
        # else: log.debug(f"Sec Row {row_idx} CUSIP {cusip}: No SecurityType ID provided.") # Optional: log missing type

        int_sched_code_raw = raw_data.get('interest_schedule_code')
        interest_schedule_instance = None
        if int_sched_code_raw:
            int_sched_code = str(int_sched_code_raw).strip()
            interest_schedule_instance = interest_schedules.get(int_sched_code)
            if not interest_schedule_instance:
                 log.warning(f"Sec Row {row_idx} CUSIP {cusip}: InterestSchedule code '{int_sched_code}' not found in DB. Setting schedule to NULL.")
        # else: log.debug(f"Sec Row {row_idx} CUSIP {cusip}: No InterestSchedule code provided.") # Optional: log missing schedule

        # Date Cleaning
        maturity_date = clean_date(raw_data.get('maturity_date'))
        issue_date = clean_date(raw_data.get('issue_date'))
        rate_effective_date = clean_date(raw_data.get('rate_effective_date'))
        call_date = clean_date(raw_data.get('call_date')) # Optional call date

        # Validation: Dates required, mat_dt > issue_dt
        if not maturity_date: log.warning(f"Sec Row {row_idx} CUSIP {cusip}: Missing maturity date. Skipping row."); skipped_rows += 1; continue
        if not issue_date: log.warning(f"Sec Row {row_idx} CUSIP {cusip}: Missing issue date. Skipping row."); skipped_rows += 1; continue
        if maturity_date <= issue_date:
            log.warning(f"Sec Row {row_idx} CUSIP {cusip}: Maturity date ({maturity_date}) not after issue date ({issue_date}). Skipping row.")
            skipped_rows += 1; continue

        # Rate Calculation
        base_rate = clean_decimal(raw_data.get('base_rate'), decimal_places=8) # Allow negative
        secondary_rate = clean_decimal(raw_data.get('secondary_rate'), decimal_places=8) # Allow negative
        effective_coupon = secondary_rate if secondary_rate is not None else base_rate
        # Ensure effective_coupon is not None if base_rate was required (assuming 'rate' column is required)
        # Allow coupon to be None (e.g., Zero Coupon Bonds) - Model allows null=True
        # if effective_coupon is None:
        #      log.warning(f"Sec Row {row_idx} CUSIP {cusip}: Effective coupon rate could not be determined (rate='{raw_data.get('base_rate')}', secrate_rate='{raw_data.get('secondary_rate')}'). Skipping row.")
        #      skipped_rows += 1; continue

        # Boolean / Choice Cleaning
        tax_code = str(raw_data.get('tax_code', '')).strip().lower()
        if tax_code not in ['e', 't']:
             log.warning(f"Sec Row {row_idx} CUSIP {cusip}: Invalid tax_cd '{raw_data.get('tax_code')}'. Skipping row.")
             skipped_rows += 1; continue

        int_calc_code = str(raw_data.get('interest_calc_code', '')).strip().lower()
        # Assuming model choices are 'a', 'c', 'h'
        if int_calc_code not in ['a', 'c', 'h']:
             log.warning(f"Sec Row {row_idx} CUSIP {cusip}: Invalid int_calc_cd '{raw_data.get('interest_calc_code')}'. Skipping row.")
             skipped_rows += 1; continue

        prin_paydown_flag_raw = raw_data.get('prin_paydown_flag')
        allows_paydown = clean_boolean_from_char(prin_paydown_flag_raw)
        if allows_paydown is None: # Check if 'y' or 'n' was provided
             log.warning(f"Sec Row {row_idx} CUSIP {cusip}: Invalid prin_paydown value '{prin_paydown_flag_raw}'. Skipping row.")
             skipped_rows += 1; continue

        # Factor Logic
        factor_from_excel = clean_decimal(raw_data.get('factor_from_excel'), decimal_places=10) # Allow negative/ > 1 initially
        factor = Decimal('1.0') # Default factor
        if allows_paydown:
            if factor_from_excel is not None:
                 factor = factor_from_excel
            else:
                 log.warning(f"Sec Row {row_idx} CUSIP {cusip}: prin_paydown is 'y' but factor is missing. Using factor=1.0.")
        elif factor_from_excel is not None and factor_from_excel != Decimal('1.0'):
             log.warning(f"Sec Row {row_idx} CUSIP {cusip}: prin_paydown is 'n' but factor '{factor_from_excel}' provided. Ignoring Excel factor, using 1.0.")

        # Integer Cleaning
        interest_day = clean_decimal(raw_data.get('interest_day')) # Clean as decimal first
        if interest_day is None or not (1 <= interest_day <= 31):
            log.warning(f"Sec Row {row_idx} CUSIP {cusip}: Invalid int_day '{raw_data.get('interest_day')}'. Skipping row.")
            skipped_rows += 1; continue
        interest_day = int(interest_day) # Convert to int after validation

        payments_per_year = clean_decimal(raw_data.get('payments_per_year'))
        if payments_per_year is None or payments_per_year <= 0:
            log.warning(f"Sec Row {row_idx} CUSIP {cusip}: Invalid ppy '{raw_data.get('payments_per_year')}'. Skipping row.")
            skipped_rows += 1; continue
        payments_per_year = int(payments_per_year)

        payment_delay_days = clean_decimal(raw_data.get('payment_delay_days'), non_negative=True)
        if payment_delay_days is None:
            log.warning(f"Sec Row {row_idx} CUSIP {cusip}: Invalid pmt_delay '{raw_data.get('payment_delay_days')}'. Skipping row.")
            skipped_rows += 1; continue
        payment_delay_days = int(payment_delay_days)

        # Optional Fields Cleaning
        description = str(raw_data.get('description', '')).strip()
        if not description: log.warning(f"Sec Row {row_idx} CUSIP {cusip}: Missing description. Skipping row."); skipped_rows += 1; continue # Description is required

        issuer_name = str(raw_data.get('issuer_name', '')).strip() or None
        currency = str(raw_data.get('currency', '')).strip().upper() or 'USD'
        callable_flag_excel = raw_data.get('callable_flag_excel')
        # Set callable_flag based on Excel OR presence of call_date
        callable_flag = clean_boolean_from_char(callable_flag_excel) if callable_flag_excel is not None else (call_date is not None)

        moody_rating = str(raw_data.get('moody_rating', '')).strip() or None
        sp_rating = str(raw_data.get('sp_rating', '')).strip() or None
        fitch_rating = str(raw_data.get('fitch_rating', '')).strip() or None
        sector = str(raw_data.get('sector', '')).strip() or None
        state_of_issuer = str(raw_data.get('state_of_issuer', '')).strip().upper() or None
        wal = clean_decimal(raw_data.get('wal'), decimal_places=3, non_negative=True) # Optional WAL

        # *** CLEAN CPR FIELD ***
        # Use decimal_places=5 as defined in the model
        cpr = clean_decimal(raw_data.get('cpr'), decimal_places=5)
        # Add validation if needed (e.g., non-negative)
        # cpr = clean_decimal(raw_data.get('cpr'), decimal_places=5, non_negative=True)
        # -----------------------

        # Prepare defaults dictionary for update_or_create
        data_defaults = {
            'description': description,
            'issue_date': issue_date,
            'maturity_date': maturity_date,
            'security_type': security_type_instance, # Assign instance or None
            'coupon': effective_coupon,
            'secondary_rate': secondary_rate,
            'rate_effective_date': rate_effective_date,
            'tax_code': tax_code,
            'interest_schedule': interest_schedule_instance, # Assign instance or None
            'interest_day': interest_day,
            'interest_calc_code': int_calc_code,
            'payments_per_year': payments_per_year,
            'allows_paydown': allows_paydown,
            'payment_delay_days': payment_delay_days,
            'factor': factor,
            # Optional fields
            'call_date': call_date,
            'wal': wal,
            # *** ADD CPR to defaults ***
            'cpr': cpr,
            # -------------------------
            'issuer_name': issuer_name,
            'currency': currency,
            'callable_flag': callable_flag,
            'moody_rating': moody_rating,
            'sp_rating': sp_rating,
            'fitch_rating': fitch_rating,
            'sector': sector,
            'state_of_issuer': state_of_issuer,
        }
        # Remove keys where value is None IF you don't want to overwrite existing DB values with NULL
        # data_defaults = {k: v for k, v in data_defaults.items() if v is not None} # Decide on this behavior

        # --- Database Operation ---
        retries = 0; success = False
        while retries < max_retries_per_row and not success:
            try:
                with transaction.atomic():
                    security, created = Security.objects.update_or_create(
                        cusip=cusip, # Match based on CUSIP (PK)
                        defaults=data_defaults
                    )
                success = True
                if created: created_count += 1; log.debug(f"Sec Row {row_idx}: Created Security: {cusip}")
                else: updated_count += 1; log.debug(f"Sec Row {row_idx}: Updated Security: {cusip}")
            except OperationalError as e:
                if 'database is locked' in str(e).lower() and retries < max_retries_per_row - 1:
                    retries += 1; wait_time = retry_delay * (2**retries); log.warning(f"Sec Row {row_idx}: DB locked {cusip}. Retry {retries}/{max_retries_per_row-1} in {wait_time:.2f}s"); time.sleep(wait_time)
                else: log.error(f"Sec Row {row_idx}: OpError {cusip} retries {retries}: {e}"); skipped_rows += 1; break
            except IntegrityError as e: log.error(f"Sec Row {row_idx}: IntegrityError {cusip}: {e}"); skipped_rows += 1; break
            except Exception as e: log.error(f"Sec Row {row_idx}: Error {cusip}: {e}", exc_info=True); skipped_rows += 1; break

    wb.close()
    result_message = f"Imported/Updated securities from {os.path.basename(file_path)}. Created: {created_count}, Updated: {updated_count}, Skipped: {skipped_rows}."
    log.info(result_message)
    return file_path


@shared_task
def import_customers_from_excel(file_path):
    """
    Imports or updates customers from an Excel file based on cust_num.
    Maps new Excel columns to the revamped Customer model.
    (Ensure this task uses the updated Customer model fields and Salesperson FK)
    """
    log.info(f"Starting customer import/update from {file_path}")
    updated_count = 0; created_count = 0; portfolio_created_count = 0
    portfolio_marked_default_count = 0; skipped_rows = 0
    max_retries_per_row = 3; retry_delay = 0.5

    # Define expected Excel headers
    header_map = {
        'cust_num': 'customer_number', # Maps to Customer model's unique Int field
        'cust_na1': 'name',
        'city': 'city',
        'state': 'state', # Expects 2-letter code
        'slsm_id': 'salesperson_id', # Maps to Salesperson FK (by salesperson_id)
        'ip_bnk': 'portfolio_accounting_code',
        'cost_funds': 'cost_of_funds_rate', # Optional percentage
        'fed_tax_bkt': 'federal_tax_bracket_rate', # Optional percentage
        'address': 'address',
        # 'zip_code': 'zip_code', # Removed zip_code mapping
    }

    try:
        wb = openpyxl.load_workbook(filename=file_path, data_only=True, read_only=True); ws = wb.active
    except FileNotFoundError: log.error(f"Customer file not found: {file_path}"); raise
    except Exception as e: log.error(f"Error opening customer file {file_path}: {e}", exc_info=True); raise

    # Read actual headers and map to internal names
    excel_headers = [str(cell.value).strip() if cell.value else None for cell in ws[1]]
    col_idx_map = {}
    processed_excel_headers = []
    for idx, excel_header in enumerate(excel_headers):
        if excel_header:
            normalized_header = excel_header.lower().replace(' ', '_')
            internal_name = header_map.get(normalized_header) # Direct lookup
            if internal_name:
                col_idx_map[internal_name] = idx
                processed_excel_headers.append(excel_header)
            else:
                log.warning(f"Customer Import: Unrecognized header '{excel_header}' in column {idx+1}. It will be ignored.")

    log.info(f"Customer Import: Found and processed Excel headers: {processed_excel_headers}")
    log.info(f"Customer Import: Mapped internal names to column indices: {col_idx_map}")

    # Check for mandatory fields based on internal names
    # Note: zip_code is removed from mandatory check
    mandatory_internal_names = ['customer_number', 'name', 'city', 'state', 'salesperson_id', 'portfolio_accounting_code']
    missing_mandatory = [name for name in mandatory_internal_names if name not in col_idx_map]
    if missing_mandatory:
        log.error(f"Mandatory headers missing from customer import file (mapped names: {missing_mandatory}). Cannot proceed.")
        wb.close()
        raise ValueError(f"Mandatory customer headers missing: {missing_mandatory}")

    # Pre-fetch Salespersons
    salespersons = {sp.salesperson_id: sp for sp in Salesperson.objects.all()}

    # Iterate through rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw_data = {}
        for internal_name, col_idx in col_idx_map.items():
            if col_idx < len(row): raw_data[internal_name] = row[col_idx]
            else: raw_data[internal_name] = None

        # --- Data Extraction and Cleaning ---
        cust_num_raw = raw_data.get('customer_number')
        customer_number = clean_decimal(cust_num_raw) # Clean as decimal first
        if customer_number is None: log.warning(f"Cust Row {row_idx}: Skip missing or invalid cust_num '{cust_num_raw}'."); skipped_rows += 1; continue
        try:
            customer_number = int(customer_number) # Convert to int
        except (ValueError, TypeError):
             log.warning(f"Cust Row {row_idx}: Could not convert cust_num '{cust_num_raw}' to integer. Skipping."); skipped_rows += 1; continue

        name = str(raw_data.get('name', '')).strip()
        if not name: log.warning(f"Cust Row {row_idx} CustNum {customer_number}: Missing name. Skipping row."); skipped_rows += 1; continue

        city = str(raw_data.get('city', '')).strip()
        if not city: log.warning(f"Cust Row {row_idx} CustNum {customer_number}: Missing city. Skipping row."); skipped_rows += 1; continue

        state = str(raw_data.get('state', '')).strip().upper()
        if len(state) != 2: # Basic validation for 2-letter code
             log.warning(f"Cust Row {row_idx} CustNum {customer_number}: Invalid state '{raw_data.get('state')}'. Skipping row.")
             skipped_rows += 1; continue

        portfolio_acc_code = str(raw_data.get('portfolio_accounting_code', '')).strip()
        if not portfolio_acc_code: log.warning(f"Cust Row {row_idx} CustNum {customer_number}: Missing ip_bnk code. Skipping row."); skipped_rows += 1; continue

        # Salesperson Lookup
        slsm_id_raw = raw_data.get('salesperson_id')
        salesperson_instance = None
        if slsm_id_raw is not None:
            slsm_id = str(slsm_id_raw).strip()
            salesperson_instance = salespersons.get(slsm_id)
            if not salesperson_instance:
                 log.warning(f"Cust Row {row_idx} CustNum {customer_number}: Salesperson ID '{slsm_id}' not found in DB. Setting salesperson to NULL.")
        else:
             log.warning(f"Cust Row {row_idx} CustNum {customer_number}: Missing slsm_id. Skipping row.") # Salesperson is required
             skipped_rows += 1; continue


        # Optional Fields
        address = str(raw_data.get('address', '')).strip() or None
        # zip_code removed
        cost_funds = clean_decimal(raw_data.get('cost_of_funds_rate'), decimal_places=8, non_negative=True)
        fed_tax = clean_decimal(raw_data.get('federal_tax_bracket_rate'), decimal_places=8, non_negative=True)

        # Prepare defaults dictionary (zip_code removed)
        customer_defaults = {
            'name': name,
            'city': city,
            'state': state,
            'salesperson': salesperson_instance, # Assign instance or None
            'portfolio_accounting_code': portfolio_acc_code,
            'address': address,
            # 'zip_code': zip_code, # Removed
            'cost_of_funds_rate': cost_funds,
            'federal_tax_bracket_rate': fed_tax,
        }
        # Remove None values if needed, but allow overwriting with NULL for optional fields
        # customer_defaults = {k: v for k, v in customer_defaults.items() if v is not None}

        # --- Database Operation & Portfolio Handling ---
        retries = 0; success = False; customer = None; customer_created = False
        portfolio = None; portfolio_created = False; portfolio_updated = False
        while retries < max_retries_per_row and not success:
            try:
                with transaction.atomic():
                    customer, customer_created = Customer.objects.update_or_create(
                        customer_number=customer_number, # Match based on unique customer_number
                        defaults=customer_defaults
                    )
                    # Default portfolio logic (remains the same)
                    portfolio_owner_name = customer.name # Name is required now
                    portfolio_name = f"{portfolio_owner_name} - Primary Holdings"
                    portfolio_defaults = {'is_default': True, 'name': portfolio_name}
                    try:
                        # Try finding by owner and is_default=True first
                        portfolio, portfolio_created = Portfolio.objects.get_or_create(
                            owner=customer, is_default=True, defaults=portfolio_defaults
                        )
                        # If found, ensure name matches convention
                        if not portfolio_created and portfolio.name != portfolio_name:
                             portfolio.name = portfolio_name
                             portfolio.save(update_fields=['name'])
                             portfolio_updated = True # Mark as updated for logging
                    except IntegrityError:
                        # This handles the case where a non-default portfolio might exist with the target name
                        # or multiple defaults somehow exist (violates constraint, but handle defensively)
                        log.warning(f"Cust Row {row_idx}: IntegrityError finding/creating default portfolio for {customer_number}. Fallback lookup by name.")
                        # Try getting by name, ensuring it's marked default
                        portfolio, portfolio_created = Portfolio.objects.get_or_create(
                            owner=customer, name=portfolio_name, defaults=portfolio_defaults
                        )
                        if not portfolio_created and not portfolio.is_default:
                            # Found by name, but wasn't default. Mark it as default.
                            portfolio.is_default = True
                            portfolio.save(update_fields=['is_default'])
                            portfolio_updated = True # Mark as updated for logging
                    except Portfolio.MultipleObjectsReturned:
                         # This means multiple portfolios exist with is_default=True for this owner
                         log.error(f"CRITICAL: Multiple default portfolios found for customer {customer_number}. Skipping default portfolio handling for this customer.")
                         portfolio = None # Cannot determine the correct default portfolio
                         portfolio_created = False
                         portfolio_updated = False


                success = True
            except OperationalError as e:
                 if 'database is locked' in str(e).lower() and retries < max_retries_per_row -1:
                    retries += 1; wait_time = retry_delay * (2**retries); log.warning(f"Cust Row {row_idx}: DB locked {customer_number}. Retry {retries}/{max_retries_per_row-1} in {wait_time:.2f}s"); time.sleep(wait_time)
                 else: log.error(f"Cust Row {row_idx}: OpError {customer_number} retries {retries}: {e}"); skipped_rows += 1; break
            except IntegrityError as e: log.error(f"Cust Row {row_idx}: IntegrityError {customer_number}: {e}", exc_info=True); skipped_rows += 1; break
            # Catch FieldError specifically if zip_code somehow still causes issues
            except FieldError as fe:
                log.error(f"Cust Row {row_idx}: Error {customer_number}: {fe}", exc_info=True)
                skipped_rows += 1; break
            except Exception as e: log.error(f"Cust Row {row_idx}: Error {customer_number}: {e}", exc_info=True); skipped_rows += 1; break

        # Logging results (remains the same)
        if success:
             if customer_created: created_count += 1; log.debug(f"Cust Row {row_idx}: Created Customer: {customer_number}")
             else: updated_count += 1; log.debug(f"Cust Row {row_idx}: Updated Customer: {customer_number}")
             if portfolio: # Only log if portfolio handling was successful
                 if portfolio_created: portfolio_created_count += 1; log.debug(f"Cust Row {row_idx}: Created default Portfolio '{portfolio.name}' for Customer: {customer_number}")
                 elif portfolio_updated: portfolio_marked_default_count += 1; log.debug(f"Cust Row {row_idx}: Marked/Updated default Portfolio '{portfolio.name}' for Customer: {customer_number}")

    wb.close()
    result_message = (f"Imported/Updated customers from {os.path.basename(file_path)}. "
                      f"Customers Created: {created_count}, Updated: {updated_count}. "
                      f"Default Portfolios Created: {portfolio_created_count}, "
                      f"Existing Portfolios Marked Default/Updated: {portfolio_marked_default_count}, "
                      f"Skipped Rows: {skipped_rows}.")
    log.info(result_message)
    return file_path


@shared_task(bind=True, autoretry_for=(OperationalError,), retry_backoff=5, max_retries=3)
def import_holdings_from_excel(self, file_path):
    """
    Imports or updates holdings from an Excel file into the default portfolio.
    Uses external_ticket as the unique key for update/create.
    Deletes holdings from default portfolios if their external_ticket is no longer in the file.
    (Ensure this task uses the updated CustomerHolding model fields)
    """
    log.info(f"Starting holding import/update (Primary Portfolio Only) from {file_path}")
    updated_count = 0; created_count = 0; deleted_count = 0; skipped_rows = 0
    # Stores {customer_id: set(external_ticket)} processed from the file
    processed_tickets_in_default_portfolio = {}

    max_retries_per_row = 3; retry_delay = 0.5

    # Define expected Excel headers
    header_map = {
        'ticket': 'external_ticket', # Maps to Holding's unique Int field
        'cust_num': 'customer_number', # Used to find Customer -> Portfolio
        'sec_id': 'cusip', # Used to find Security
        'lc_xf1_cd': 'intention_code', # Expects 'A', 'M', 'T'
        'orig_face': 'original_face_amount',
        'set_par': 'settled_par', # Ignored as per instructions, but read if present
        'settle_dt': 'settlement_date',
        'set_price': 'settlement_price',
        'hold_duration': 'holding_duration', # Optional
        'hold_avg_life': 'holding_average_life', # Optional
        'hold_avg_life_dt': 'holding_average_life_date', # Optional
        'mkt_dt': 'market_date', # Optional
        'mkt_price': 'market_price', # Optional
        'mkt_yield': 'market_yield', # Optional
        'book_price': 'book_price',
        'book_yield': 'book_yield', # Optional
    }

    try:
        wb = openpyxl.load_workbook(filename=file_path, data_only=True, read_only=True); ws = wb.active
    except FileNotFoundError: log.error(f"Holding import file not found: {file_path}"); raise
    except Exception as e: log.error(f"Error opening holding import file {file_path}: {e}", exc_info=True); raise

    # Read actual headers and map to internal names
    excel_headers = [str(cell.value).strip() if cell.value else None for cell in ws[1]]
    col_idx_map = {}
    processed_excel_headers = []
    for idx, excel_header in enumerate(excel_headers):
        if excel_header:
            normalized_header = excel_header.lower().replace(' ', '_')
            internal_name = header_map.get(normalized_header)
            if internal_name:
                col_idx_map[internal_name] = idx
                processed_excel_headers.append(excel_header)
            else:
                log.warning(f"Holding Import: Unrecognized header '{excel_header}' in column {idx+1}. It will be ignored.")

    log.info(f"Holding Import: Found and processed Excel headers: {processed_excel_headers}")
    log.info(f"Holding Import: Mapped internal names to column indices: {col_idx_map}")

    # Check for mandatory fields based on internal names
    mandatory_internal_names = [
        'external_ticket', 'customer_number', 'cusip', 'intention_code',
        'original_face_amount', 'settlement_date', 'settlement_price', 'book_price'
    ]
    missing_mandatory = [name for name in mandatory_internal_names if name not in col_idx_map]
    if missing_mandatory:
        log.error(f"Mandatory headers missing from holding import file (mapped names: {missing_mandatory}). Cannot proceed.")
        wb.close()
        raise ValueError(f"Mandatory holding headers missing: {missing_mandatory}")

    # --- Caching ---
    customer_cache = {} # {cust_num: customer_instance}
    security_cache = {} # {cusip: security_instance}
    default_portfolio_cache = {} # {customer_id: portfolio_instance_or_None}

    # --- First Pass: Process rows, update/create holdings in DEFAULT portfolio ---
    log.info("Holdings Pass 1: Updating/Creating holdings in default portfolios...")
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw_data = {}
        for internal_name, col_idx in col_idx_map.items():
            if col_idx < len(row): raw_data[internal_name] = row[col_idx]
            else: raw_data[internal_name] = None

        # --- Data Extraction and Cleaning ---
        ext_ticket_raw = raw_data.get('external_ticket')
        external_ticket = clean_decimal(ext_ticket_raw)
        if external_ticket is None: log.warning(f"Hold Row {row_idx}: Skip missing or invalid ticket '{ext_ticket_raw}'."); skipped_rows += 1; continue
        try:
            external_ticket = int(external_ticket)
        except (ValueError, TypeError):
            log.warning(f"Hold Row {row_idx}: Could not convert ticket '{ext_ticket_raw}' to integer. Skipping."); skipped_rows += 1; continue

        cust_num_raw = raw_data.get('customer_number')
        customer_number = clean_decimal(cust_num_raw)
        if customer_number is None: log.warning(f"Hold Row {row_idx} Ticket {external_ticket}: Skip missing or invalid cust_num '{cust_num_raw}'."); skipped_rows += 1; continue
        try:
            customer_number = int(customer_number)
        except (ValueError, TypeError):
            log.warning(f"Hold Row {row_idx} Ticket {external_ticket}: Could not convert cust_num '{cust_num_raw}' to integer. Skipping."); skipped_rows += 1; continue

        cusip_raw = raw_data.get('cusip')
        if not cusip_raw: log.warning(f"Hold Row {row_idx} Ticket {external_ticket}: Skip missing cusip."); skipped_rows += 1; continue
        cusip = str(cusip_raw).strip().upper()
        # Basic CUSIP validation
        if len(cusip) != 9 or not cusip.isalnum():
            log.warning(f"Hold Row {row_idx} Ticket {external_ticket}: Invalid CUSIP format '{cusip_raw}'. Skipping row.")
            skipped_rows += 1; continue

        # --- Get Customer (use cache) ---
        customer = customer_cache.get(customer_number)
        if not customer:
            try:
                customer = Customer.objects.get(customer_number=customer_number)
                customer_cache[customer_number] = customer
            except Customer.DoesNotExist: log.warning(f"Hold Row {row_idx} Ticket {external_ticket}: Skip unknown customer_number {customer_number}"); skipped_rows += 1; continue
            except Exception as e: log.error(f"Hold Row {row_idx} Ticket {external_ticket}: Error fetch customer {customer_number}: {e}"); skipped_rows += 1; continue

        # --- Get Security (use cache) ---
        security = security_cache.get(cusip)
        if not security:
            try:
                security = Security.objects.get(cusip=cusip)
                security_cache[cusip] = security
            except Security.DoesNotExist: log.warning(f"Hold Row {row_idx} Ticket {external_ticket}: Skip unknown cusip {cusip}"); skipped_rows += 1; continue
            except Exception as e: log.error(f"Hold Row {row_idx} Ticket {external_ticket}: Error fetch security {cusip}: {e}"); skipped_rows += 1; continue

        # --- Get Default Portfolio (use cache) ---
        default_portfolio = default_portfolio_cache.get(customer.id)
        if customer.id not in default_portfolio_cache: # Check cache first
            try:
                default_portfolio = Portfolio.objects.get(owner=customer, is_default=True)
                default_portfolio_cache[customer.id] = default_portfolio
            except Portfolio.DoesNotExist: log.warning(f"Hold Row {row_idx} Ticket {external_ticket}: Default portfolio not found for customer {customer_number}. Skipping."); default_portfolio_cache[customer.id] = None; skipped_rows += 1; continue
            except Portfolio.MultipleObjectsReturned: log.error(f"Hold Row {row_idx} Ticket {external_ticket}: CRITICAL - Multiple default portfolios found for customer {customer_number}. Skipping."); default_portfolio_cache[customer.id] = None; skipped_rows += 1; continue
            except Exception as e: log.error(f"Hold Row {row_idx} Ticket {external_ticket}: Error fetch default portfolio for {customer_number}: {e}"); default_portfolio_cache[customer.id] = None; skipped_rows += 1; continue
        if default_portfolio is None: skipped_rows += 1; continue # Skip if default portfolio known not to exist

        # --- Clean Remaining Fields ---
        intention_code = str(raw_data.get('intention_code', '')).strip().upper()
        if intention_code not in ['A', 'M', 'T']: log.warning(f"Hold Row {row_idx} Ticket {external_ticket}: Invalid lc_xf1_cd '{raw_data.get('intention_code')}'. Skipping row."); skipped_rows += 1; continue

        orig_face = clean_decimal(raw_data.get('original_face_amount'), decimal_places=8, non_negative=True)
        if orig_face is None: log.warning(f"Hold Row {row_idx} Ticket {external_ticket}: Invalid orig_face '{raw_data.get('original_face_amount')}'. Skipping row."); skipped_rows += 1; continue

        settle_dt = clean_date(raw_data.get('settlement_date'))
        if settle_dt is None: log.warning(f"Hold Row {row_idx} Ticket {external_ticket}: Invalid settle_dt '{raw_data.get('settlement_date')}'. Skipping row."); skipped_rows += 1; continue

        set_price = clean_decimal(raw_data.get('settlement_price'), decimal_places=8, non_negative=True)
        if set_price is None: log.warning(f"Hold Row {row_idx} Ticket {external_ticket}: Invalid set_price '{raw_data.get('settlement_price')}'. Skipping row."); skipped_rows += 1; continue

        book_price = clean_decimal(raw_data.get('book_price'), decimal_places=8, non_negative=True)
        if book_price is None: log.warning(f"Hold Row {row_idx} Ticket {external_ticket}: Invalid book_price '{raw_data.get('book_price')}'. Skipping row."); skipped_rows += 1; continue

        # Optional fields
        book_yield = clean_decimal(raw_data.get('book_yield'), decimal_places=8, non_negative=True)
        hold_duration = clean_decimal(raw_data.get('holding_duration'), decimal_places=8, non_negative=True)
        hold_avg_life = clean_decimal(raw_data.get('holding_average_life'), decimal_places=8, non_negative=True)
        hold_avg_life_dt = clean_date(raw_data.get('holding_average_life_date'))
        mkt_dt = clean_date(raw_data.get('market_date'))
        mkt_price = clean_decimal(raw_data.get('market_price'), decimal_places=8, non_negative=True)
        mkt_yield = clean_decimal(raw_data.get('market_yield'), decimal_places=8, non_negative=True)

        # --- Prepare holding data defaults ---
        # Include external_ticket in defaults now
        holding_defaults = {
            # 'external_ticket': external_ticket, # Don't update external_ticket, use it for lookup only
            'portfolio': default_portfolio, # Link to the specific default portfolio
            'security': security,
            'intention_code': intention_code,
            'original_face_amount': orig_face,
            'settlement_date': settle_dt,
            'settlement_price': set_price,
            'book_price': book_price,
            'book_yield': book_yield,
            'holding_duration': hold_duration,
            'holding_average_life': hold_avg_life,
            'holding_average_life_date': hold_avg_life_dt,
            'market_date': mkt_dt,
            'market_price': mkt_price,
            'market_yield': mkt_yield,
        }
        # Remove None values if needed
        # holding_defaults = {k: v for k, v in holding_defaults.items() if v is not None}

        # --- Database Operation (Update/Create based on external_ticket) ---
        retries = 0; success = False
        while retries < max_retries_per_row and not success:
            try:
                with transaction.atomic():
                    # Use external_ticket as the unique key for finding/updating holdings
                    holding, created = CustomerHolding.objects.update_or_create(
                        external_ticket=external_ticket, # Match on external ticket
                        defaults=holding_defaults # Update all other fields
                    )
                success = True
                # Track processed external tickets *for this customer's default portfolio*
                if customer.id not in processed_tickets_in_default_portfolio:
                     processed_tickets_in_default_portfolio[customer.id] = set()
                # Still track the external ticket from the file for the deletion phase
                processed_tickets_in_default_portfolio[customer.id].add(external_ticket)

                if created: created_count += 1; log.debug(f"Hold Row {row_idx}: Created Holding Ticket {external_ticket} for Sec {cusip} in '{default_portfolio.name}'")
                else: updated_count += 1; log.debug(f"Hold Row {row_idx}: Updated Holding Ticket {external_ticket} for Sec {cusip} in '{default_portfolio.name}'")

            except OperationalError as e:
                 # Let celery handle retry based on task decorator
                 log.warning(f"Hold Row {row_idx}: DB locked for holding ticket {external_ticket}. Celery will retry...")
                 self.retry(exc=e) # Trigger Celery's retry mechanism
                 # Note: If retry fails max times, task fails. Need error handling strategy.
                 # For now, just log and let it fail if lock persists.
                 log.error(f"Hold Row {row_idx}: DB lock persisted for holding ticket {external_ticket}. Task may fail.")
                 skipped_rows += 1; break # Break local loop after triggering retry
            except IntegrityError as e:
                # This might catch UNIQUE constraint on external_ticket if it's somehow duplicated
                # or if portfolio+security constraint was re-added accidentally.
                log.error(f"Hold Row {row_idx}: IntegrityError on holding ticket {external_ticket} (Port: {default_portfolio.id}, Sec: {cusip}): {e}", exc_info=True)
                skipped_rows += 1; break
            except Exception as e:
                log.error(f"Hold Row {row_idx}: Unexpected error on holding ticket {external_ticket}: {e}", exc_info=True)
                skipped_rows += 1; break

    # --- Second Pass: Delete obsolete holdings from relevant DEFAULT portfolios ---
    log.info("Holdings Pass 2: Deleting obsolete holdings from default portfolios...")
    all_processed_tickets = set()
    for ticket_set in processed_tickets_in_default_portfolio.values():
        all_processed_tickets.update(ticket_set)

    # Find all customers who had holdings processed in this run
    customers_to_check_ids = list(processed_tickets_in_default_portfolio.keys())
    if customers_to_check_ids:
        # Get all default portfolios for these customers
        relevant_default_portfolios = Portfolio.objects.filter(
            owner_id__in=customers_to_check_ids,
            is_default=True
        )
        # Find holdings in these default portfolios whose external_ticket is NOT in the set we just processed
        obsolete_holdings_qs = CustomerHolding.objects.filter(
            portfolio__in=relevant_default_portfolios
        ).exclude(
            external_ticket__in=all_processed_tickets
        )

        try:
            # Perform deletion in bulk
            deleted_count, deleted_details = obsolete_holdings_qs.delete()
            if deleted_count > 0:
                log.info(f"Holdings Deletion: Deleted {deleted_count} obsolete holdings from relevant default portfolios. Details: {deleted_details}")
        except Exception as e:
             log.error(f"Holdings Deletion: Error deleting obsolete holdings: {e}", exc_info=True)
             # Decide on error handling - stop task? Log and continue?
    else:
         log.info("Holdings Deletion: No customers had holdings processed, skipping deletion phase.")


    wb.close()
    result_message = (f"Processed holdings (Primary Portfolio Only) from {os.path.basename(file_path)}. "
                      f"Created: {created_count}, Updated: {updated_count}, Deleted: {deleted_count}, Skipped Rows: {skipped_rows}.")
    log.info(result_message)
    return file_path


# --- UPDATED Muni Offering Import Task ---
@shared_task(bind=True, max_retries=1) # Keep max_retries=1 for muni? Or allow more?
def import_muni_offerings_from_excel(self, file_path):
    """
    Imports or updates municipal offerings from an Excel file based on CUSIP.
    Includes enhanced cleaning for CUSIP and dates.
    """
    log.info(f"Starting municipal offering import/update from {file_path}")
    created_count = 0; updated_count = 0; skipped_rows = 0; deleted_count = 0

    try:
        wb = openpyxl.load_workbook(filename=file_path, data_only=True, read_only=True); ws = wb.active
    except FileNotFoundError: log.error(f"Municipal offering import file not found: {file_path}"); raise
    except Exception as e: log.error(f"Error opening municipal offering import file {file_path}: {e}", exc_info=True); raise

    # --- Pre-import Deletion ---
    try:
        with transaction.atomic():
            log.warning("Deleting ALL existing MunicipalOffering records before import...")
            deleted_count, _ = MunicipalOffering.objects.all().delete(); log.info(f"Deleted {deleted_count} existing MunicipalOffering records.")
    except Exception as e: log.error(f"Error during pre-import deletion of MunicipalOfferings: {e}", exc_info=True); wb.close(); raise

    # --- Header Mapping ---
    headers = [str(cell.value).lower().strip().replace(' ', '_').replace('&', 'and') if cell.value else None for cell in ws[1]]
    headers = [h for h in headers if h]; log.info(f"Found muni offering headers: {headers}")
    # Define mapping from expected Excel headers to model fields
    header_map = {
        'cusip': 'cusip', 'amount': 'amount', 'description': 'description', 'coupon': 'coupon',
        'maturity': 'maturity_date', 'yield': 'yield_rate', 'price': 'price', 'moody': 'moody_rating',
        's_and_p': 'sp_rating', 'call_date': 'call_date', 'call_price': 'call_price', 'state': 'state',
        'insurance': 'insurance',
    }
    # Check for mandatory CUSIP header
    if 'cusip' not in headers: log.error("Mandatory header 'cusip' not found."); wb.close(); raise ValueError("Mandatory header 'cusip' not found.")

    # --- Row Processing ---
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Check for completely empty rows (common in Excel)
        if all(cell is None for cell in row):
            log.debug(f"Muni Row {row_idx}: Skipping empty row.")
            continue

        # Ensure row has enough columns to match headers found
        if len(row) < len(headers):
             log.warning(f"Muni Row {row_idx}: Row has fewer columns ({len(row)}) than headers ({len(headers)}). Skipping.")
             skipped_rows += 1; continue

        raw_data = dict(zip(headers, row))

        # --- CUSIP Cleaning and Validation ---
        cusip_raw = raw_data.get('cusip')
        if cusip_raw is None or str(cusip_raw).strip() == '':
             log.warning(f"Muni Row {row_idx}: Skipping missing CUSIP."); skipped_rows += 1; continue
        cusip = str(cusip_raw).strip().upper()
        # Stricter CUSIP validation (9 chars, alphanumeric)
        if len(cusip) != 9 or not cusip.isalnum():
            log.warning(f"Muni Row {row_idx}: Invalid CUSIP format '{cusip_raw}'. Skipping row.")
            skipped_rows += 1; continue

        # --- Data Cleaning for other fields ---
        offering_defaults = {}; skip_this_row = False
        for excel_header, model_field in header_map.items(): # Iterate through defined map
             if excel_header in raw_data:
                 raw_value = raw_data[excel_header]
                 cleaned_value = None

                 # *** Skip cleaning for CUSIP as it's already validated ***
                 if model_field == 'cusip':
                     continue # Already handled above

                 # Clean numeric fields
                 elif model_field in ['amount', 'coupon', 'yield_rate', 'price', 'call_price']:
                     cleaned_value = clean_decimal(raw_value, decimal_places=6) # Adjust precision if needed
                     # Add non-negative validation if applicable
                     # cleaned_value = clean_decimal(raw_value, decimal_places=6, non_negative=True)

                 # Clean date fields
                 elif model_field in ['maturity_date', 'call_date']:
                     cleaned_value = clean_date(raw_value) # Use updated clean_date
                     # Add validation: Skip if required date (e.g., maturity) is missing/invalid
                     if model_field == 'maturity_date' and cleaned_value is None and raw_value is not None:
                          log.warning(f"Muni Row {row_idx} CUSIP {cusip}: Invalid or missing maturity date '{raw_value}'. Skipping row.")
                          skip_this_row = True; break # Stop processing this row
                     elif cleaned_value is None and raw_value is not None:
                          # Log if cleaning failed for optional date but value was present
                          log.warning(f"Muni Row {row_idx} CUSIP {cusip}: Failed to clean date for '{model_field}' from value '{raw_value}'. Storing as NULL.")

                 # Clean state field
                 elif model_field == 'state':
                     cleaned_value = str(raw_value).strip().upper() if raw_value is not None else None
                     if cleaned_value and len(cleaned_value) != 2:
                         log.warning(f"Muni Row {row_idx} CUSIP {cusip}: Invalid state format '{raw_value}'. Storing as NULL.")
                         cleaned_value = None # Store as NULL if invalid format

                 # Handle other string fields
                 else:
                     cleaned_value = str(raw_value).strip() if raw_value is not None else None

                 # Store cleaned value if valid, otherwise keep it None
                 # We store None even if cleaning failed but value was present (logged above)
                 offering_defaults[model_field] = cleaned_value

             # else: # Log missing headers? Optional.
             #     if excel_header in header_map:
             #         log.warning(f"Muni Row {row_idx} CUSIP {cusip}: Missing expected header: {excel_header}")
             #         offering_defaults[model_field] = None # Ensure field is None if header missing

        if skip_this_row: skipped_rows += 1; continue # Skip row if mandatory field failed cleaning

        # --- Database operation ---
        try:
            with transaction.atomic():
                offering, created = MunicipalOffering.objects.update_or_create(
                    cusip=cusip, # Match on CUSIP
                    defaults=offering_defaults # Update with cleaned values
                )
            log.debug(f"Muni Row {row_idx}: {'Created' if created else 'Updated'} Offering: {cusip}")
            if created: created_count += 1
            else: updated_count += 1
        except IntegrityError as e: # Catch potential DB constraint violations
            log.error(f"Muni Row {row_idx}: IntegrityError processing {cusip}: {e}", exc_info=True)
            skipped_rows += 1
        except Exception as e:
            log.error(f"Muni Row {row_idx}: Error processing {cusip}: {e}", exc_info=True)
            skipped_rows += 1

    wb.close()
    result_message = (f"Imported/Updated municipal offerings from {os.path.basename(file_path)}. "
                      f"Deleted: {deleted_count}, Created: {created_count}, Updated: {updated_count}, Skipped: {skipped_rows}.")
    log.info(result_message)
    return result_message


# --- UPDATED import_all_from_excel Task ---
# (No changes needed here, keeps the correct sequence)
@shared_task
def import_all_from_excel():
    """
    Orchestrates the import tasks in sequence using hardcoded paths.
    Assumes standard filenames. Includes LOOKUP imports first.
    Uses immutable signatures (.si) to prevent passing results.
    """
    log.info("Scheduling chained non-destructive import from hardcoded paths...")
    base = settings.BASE_DIR / 'data' / 'imports'

    # Define expected filenames (case-insensitive matching might be good)
    filenames = {
        'salesperson': base / 'Salesperson.xlsx',
        'security_type': base / 'SecurityType.xlsx',
        'interest_schedule': base / 'InterestSchedule.xlsx',
        'security': base / 'Security.xlsx',
        'customer': base / 'Customer.xlsx',
        'holding': base / 'Holdings.xlsx',
        'muni_offering': base / 'muni_offerings.xlsx',
    }

    # Map internal names to tasks and filenames
    import_config = {
        'salesperson': (import_salespersons_from_excel, filenames['salesperson']),
        'security_type': (import_security_types_from_excel, filenames['security_type']),
        'interest_schedule': (import_interest_schedules_from_excel, filenames['interest_schedule']),
        'security': (import_securities_from_excel, filenames['security']),
        'customer': (import_customers_from_excel, filenames['customer']),
        'holding': (import_holdings_from_excel, filenames['holding']),
        'muni_offering': (import_muni_offerings_from_excel, filenames['muni_offering']),
    }

    # Define the desired import order
    import_order = [
        'salesperson', 'security_type', 'interest_schedule', # Lookups First
        'security',
        'customer',
        'holding',
        'muni_offering' # Muni offerings last (or adjust if needed)
    ]

    task_list = []
    files_ok = True
    found_files_count = 0

    # Build the task list in the correct order, checking file existence
    for import_key in import_order:
        task_func, file_path = import_config[import_key]
        if file_path.exists():
            task_list.append(task_func.si(str(file_path)))
            log.info(f"Chained Import: Found '{os.path.basename(file_path)}', adding task '{task_func.__name__}'.")
            found_files_count += 1
        else:
            # Log error for mandatory files, warning for optional (like muni)
            if import_key in ['salesperson', 'security_type', 'interest_schedule', 'security', 'customer', 'holding']:
                log.error(f"Chained Import ERROR: Mandatory file not found: {file_path}")
                files_ok = False # Mark overall status as problematic
            else:
                log.warning(f"Chained Import: Optional file not found: {file_path}. Skipping {task_func.__name__}.")

    if not task_list:
        result_message = "Chained Import Error: No import tasks could be added (no files found or critical files missing)."
        log.error(result_message)
        return result_message
    elif not files_ok:
         log.error("Chained Import Error: One or more mandatory import files were missing. Chain may be incomplete.")
         # Proceed with the chain anyway, but log the error clearly

    # Create and run the chain
    import_chain = chain(task_list)
    import_chain.apply_async()

    result_message = f"Scheduled chained import tasks ({len(task_list)} tasks from {found_files_count} files). Mandatory File Status OK: {files_ok}."
    log.info(result_message)
    return result_message


# --- Email Tasks (No changes needed in this step) ---

@shared_task(bind=True, autoretry_for=(Exception,), retry_backoff=True, max_retries=3)
def send_salesperson_interest_email(self, salesperson_email, salesperson_name, customer_name, customer_number, selected_bonds):
    """ Sends email notification about customer's selling interest. """
    log.info(f"Task send_salesperson_interest_email started for salesperson: {salesperson_email}, customer: {customer_number}")
    subject = f"Interest in Selling Bonds - Customer {customer_name} ({customer_number})"
    greeting_name = salesperson_name if salesperson_name else "Salesperson"
    customer_display = f"{customer_name} ({customer_number})" if customer_name else f"Customer {customer_number}"
    bond_lines = []
    for bond in selected_bonds:
        try: par_decimal = Decimal(bond['par']); par_formatted = f"{par_decimal:,.2f}"
        except (InvalidOperation, TypeError, ValueError): par_formatted = bond['par'] # Keep original string if conversion fails
        bond_lines.append(f"  - CUSIP: {bond['cusip']}, Par: {par_formatted}")
    bonds_list_str = "\n".join(bond_lines)
    body = f"""Dear {greeting_name},\n\nOur client, {customer_display}, has indicated interest in selling the following bonds held in their portfolio:\n\n{bonds_list_str}\n\nPlease follow up with them regarding this interest.\n\nThanks,\nPortfolio Analyzer System\n"""

    try:
        send_mail(subject, body, settings.DEFAULT_FROM_EMAIL, [salesperson_email], fail_silently=False)
        log.info(f"Successfully sent interest email to {salesperson_email} for customer {customer_number}")
        return f"Email sent successfully to {salesperson_email}"
    except Exception as e:
        log.error(f"Failed to send interest email to {salesperson_email} for customer {customer_number}. Error: {e}", exc_info=True)
        # Retry logic handled by task decorator
        raise self.retry(exc=e)


@shared_task(bind=True, autoretry_for=(Exception,), retry_backoff=True, max_retries=3)
def send_salesperson_muni_buy_interest_email(self, salesperson_email, salesperson_name, customer_name, customer_number, selected_offerings):
    """ Sends email notification about customer's interest in buying municipal offerings. """
    log.info(f"Task send_salesperson_muni_buy_interest_email started for salesperson: {salesperson_email}, customer: {customer_number}")
    subject = f"Interest in Buying Municipal Offerings - Customer {customer_name} ({customer_number})"
    greeting_name = salesperson_name if salesperson_name else "Salesperson"
    customer_display = f"{customer_name} ({customer_number})" if customer_name else f"Customer {customer_number}"
    offering_lines = []
    for offering in selected_offerings: offering_lines.append(f"  - CUSIP: {offering['cusip']} ({offering.get('description', 'N/A')})")
    offerings_list_str = "\n".join(offering_lines)
    body = f"""Dear {greeting_name},\n\nOur client, {customer_display}, has indicated interest in potentially buying the following municipal bond offerings:\n\n{offerings_list_str}\n\nPlease follow up with them regarding this interest and availability.\n\nThanks,\nPortfolio Analyzer System\n"""

    try:
        send_mail(subject, body, settings.DEFAULT_FROM_EMAIL, [salesperson_email], fail_silently=False)
        log.info(f"Successfully sent muni buy interest email to {salesperson_email} for customer {customer_number}")
        return f"Email sent successfully to {salesperson_email}"
    except Exception as e:
        log.error(f"Failed to send muni buy interest email to {salesperson_email} for customer {customer_number}. Error: {e}", exc_info=True)
        raise self.retry(exc=e)
